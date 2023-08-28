import tushare as ts
import pandas as pd
import matplotlib.pyplot as plt
import json
import numpy as np
from matplotlib.pylab import date2num
import datetime
import mpl_finance as mpf
import pyecharts
from pyecharts import options as opts
from pyecharts.commons.utils import JsCode
from pyecharts.charts import *
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
 
 
 
# 设置Token
ts.set_token('35161dbcb4b76f532f815a1bee71c10a1b539c67a646462d9600616b') #请填入自己的Token
 
# 初始化接口
ts_api = ts.pro_api()
print("Tushare 接口调用正常，版本号:" + ts.__version__ +"\n")
print("pyecharts数据交互 接口调用正常，版本号:{}".format(pyecharts.__version__)+"\n")
 
 
class RobotTrade():
    id = None
    buyTime = None
    buyPrice = None
    sellTime = None
    sellPrice = None
    profit = None
 
    def setId(self,id):
        self.id = id
 
    def buy(self,time,price):
        self.buyTime = time
        self.buyPrice = price
 
    def sell(self,time,price):
        self.sellTime = time
        self.sellPrice = price
        self.profit = round((self.sellPrice - self.buyPrice)/self.buyPrice,4)
 
    def Out(self):
        a = self.buyTime
        b = self.sellTime
        c = self.buyPrice
        d = self.sellPrice
        e = self.profit
        f = self.id
        return [a,b,c,d,e,f]
 
class Stock:
    global returnList
    isTodayBuy = ""
    df = pd.DataFrame([0, 0])
    code = ""
    start = ""
    end = ""
    name = ""
    list_BuyorSell = []
 
    # 设置股票代码
    def set_code(self,code):
 
        # stock_list = ts_api.stock_basic(exchange='', list_status='L',
        #                                 fields='ts_code,symbol,name,area,industry,market,list_date')
        # print(stock_list)


        self.code = code+".sz"
        stock_list = pd.read_csv('all/all-code.csv')
        a = stock_list["ts_code"].tolist()

        for i in range(len(a)):
            if a[i][0:6] == code:
                self.code = a[i]
                self.name = stock_list.loc[i, "name"]
        #self.name = "***"
            # 请求股票列表
 
    # 获取日线交易数据
    def get_daily_trade(self):  # 股票代码，开始日期20120101，结束日期20180101
 
        # print("正在获取交易数据：\n    --Start Download")
        if self.start == "":
            self.start = input("请输入交易开始日期：")
            self.end = input("请输入交易结束日期：")
 
        # 请求日线数据
        # print(self.start)
        df = ts_api.daily(ts_code = self.code, start_date = self.start, end_date = self.end)
        # print(df)
 
        # 按交易日升序排序
        df.sort_values(by=["trade_date"], inplace=True)
        df.reset_index(drop=True, inplace=True)
        self.df = df
 
        # print("已获取{} {}，交易数据，共计{}条交易数据\n".format(self.code,self.name, len(self.df)))
 
    # 指数移动均线工具
    def get_MA_EMA(self, n):
        df = self.df
        newname = "EMA" + str(n)
        df[newname] = 0  # 初始化n天指数移动均线
        for i in range(n - 1, len(df)):  # 循环第n个数据开始，为之后的每个数据计算EMA
            list_temp = []
            list_ema = []
            for d in range(n):  # 添加数据
                list_temp.append(df.loc[i - (n - d) + 1, "close"])
                list_ema.append(0)
            for d in range(n):  # 进行计算
                if d == 0:
                    list_ema[d] = list_temp[d]
                else:
                    list_ema[d] = (2 * list_temp[d] + (n - 1) * list_ema[d - 1]) / (n + 1)
            df.loc[i, newname] = list_ema[n - 1]
        # print("    --已计算指数移动均线EMA{}".format(n))
        self.df = df
 
    # 移动均线计算
    def get_MA(self, n):
        df = self.df
        newname = "MA" + str(n)
        df[newname] = None  # 初始化n天指数移动均线
        for i in range(n-1,len(df)):
            list_temp = []
            sum = 0
            avg = 0
            for d in range(n):
                sum = sum + df.loc[ i-(n-d)+1 ,"close"]
            avg = round(sum/n,2)
            df.loc[i, newname] = avg
        # print("    --已计算移动均线MA{}".format(n))
 
    # 移动均线计算
    def get_Vol_MA(self, n):
        df = self.df
        newname = "VolMA" + str(n)
        df[newname] = None  # 初始化n天指数移动均线
        for i in range(n-1,len(df)):
            list_temp = []
            sum = 0
            avg = 0
            for d in range(n):
                sum = sum + df.loc[ i-(n-d)+1 ,"vol"]
            avg = round(sum/n,2)
            df.loc[i, newname] = avg
        # print("    --已计算成交量移动均线MA{}".format(n))
 
    # Macd的快速线DIF计算
    def get_Macd_DIF(self, n1, n2):
        df = self.df
        newname = "Macd_DIF"
        df[newname] = 0  # 初始化n天指数移动均线
        for i in range(n2,len(df)):
            df.loc[i, "Macd_DIF"] = df.loc[i, "EMA" + str(n1)] - df.loc[i, "EMA" + str(n2)]
        # print("    --已计算MACD 快速线DIF")
        self.df = df
 
    # Macd的慢速线DEA计算
    def get_Macd_DEA(self, n):
        df = self.df
        newname = "Macd_DEA"
        df[newname] = 0  # 初始化n天指数移动均线
        for i in range(n - 1, len(df)):  # 循环第n个数据开始，为之后的每个数据计算EMA
            list_temp = []
            list_dif = []
            for d in range(n):  # 添加数据
                list_temp.append(df.loc[i - (n - d) + 1, "Macd_DIF"])
                list_dif.append(0)
            for d in range(n):  # 进行计算
                if d == 0:
                    list_dif[d] = list_temp[d]
                else:
                    list_dif[d] = (2 * list_temp[d] + (n - 1) * list_dif[d - 1]) / (n + 1)
            df.loc[i, newname] = list_dif[n - 1]
        # print("    --已计算MACD 慢速线DEA")
        self.df = df
 
    # 计算MACD柱状线
    def get_Macd_Bar(self,n):
        df = self.df
        for i in range(n,len(df)):
            df.loc[i,"MACD_Bar"] = 2 * ( df.loc[i, "Macd_DIF"] - df.loc[i, "Macd_DEA"])
        # print("    --已计算MACD 柱状线Bar")
        self.df = df
 
    # Macd计算
    def get_Macd(self, a=12, b=26, c=9):
 
        # print("正在计算MACD指标：")
 
        # 计算EMA12、EMA26 数据
        self.get_MA_EMA(a)
        self.get_MA_EMA(b)
 
        # 计算三条曲线
        self.get_Macd_DIF(a,b)
        self.get_Macd_DEA(c)
        self.get_Macd_Bar(b)
 
        # print("    --已完成MACD指标的计算")
 
    # 交易策略
    def set_Trading_Strategy(self):
        df = self.df
        list_BuyorSell = []
        count = 0
        temp = 0 # 买入后的世界线进行
 
        for i in range(10,len(df)): # 偷懒从第十个交易日开始
            i = i + temp
            if i >=len(df):
                break
            lowThanma5 = 0 # 收盘价低于5日均线的天数
            Is_skip = False # 是否条件不满足
            Is_selled = False
 
 
            # 条件0：买点的前2天收盘价均小于5日均线
            for day in range(1,2):
                if df.loc[i-day,"MA5"] < df.loc[i-day,"close"]:
                    Is_skip = True
            if Is_skip==True:
                continue
 
            # 条件1：当天收盘价大于5日线
            if df.loc[i,"close"] < df.loc[i,"MA5"]:
                Is_skip = True
            if Is_skip==True:
                continue
 
            # # 条件2：当天开盘价低于5日线
            # if df.loc[i, "open"] > df.loc[i, "MA5"]:
            #     Is_skip = True
            # if Is_skip == True:
            #     continue
 
            # 条件3: 当日增幅超过3%
            if ((df.loc[i,"close"]-df.loc[i,"open"]) / df.loc[i,"open"] ) <= 0.03:
                Is_skip = True
            if Is_skip==True:
                continue
 
            # 条件3:...
 
            if Is_skip==True:
                continue
            # 买入条件均满足后买入
            r = RobotTrade() # 设置交易机器人id
            r.setId(count)
 
            if i+1 == len(df): # 是否今天买入
                self.isTodayBuy = "下个交易日开盘立刻买入"
                break
            r.buy(i+1,df.loc[i+1,"open"]) # 机器人后一天买入
            have_price = df.loc[i+1,"open"] # 第二天开盘立刻买入
 
            # 一但卖出条件成熟，机器人执行卖出
            for day in range(1,31): # 最大持股时间30日
 
                if i+1+day >= len(df) - 1: # 检查数组是否越界
                    r.sell(len(df) - 1, r.buyPrice)
                    list_BuyorSell.append(r.Out())
                    break
 
                # 单次操作亏损达到-0.0666时卖出
                if (df.loc[i+1+day,"close"]-r.buyPrice)/r.buyPrice <= -0.0666:
                    r.sell(i + 1 + day, r.buyPrice*0.9334)
                    list_BuyorSell.append(r.Out())
                    Is_selled = True
                    break
 
                # 单次操作盈利达到+0.2888时卖出
                if (df.loc[i + 1 + day, "high"] - r.buyPrice) / r.buyPrice >= 0.1888:
                    r.sell(i + 1 + day, r.buyPrice * 1.1888)
                    list_BuyorSell.append(r.Out())
                    Is_selled = True
                    break
 
                # 5日均线小于10日均线的0.98 卖出
                if df.loc[i+1+day,"MA10"]*0.98 > df.loc[i+1+day,"MA5"]:
                    r.sell(i+1+day,df.loc[i+1+day,"close"])
                    list_BuyorSell.append(r.Out())
                    Is_selled = True
                    break
 
                # 连续3天收盘价小于5日均线 且 5日线小于10日线 卖出
                if df.loc[i + 1 + day, "MA5"] > df.loc[i + 1 + day, "close"] and df.loc[i + 1 + day, "MA5"] < \
                        df.loc[i + 1 + day, "MA10"]:
                    lowThanma5 = lowThanma5 + 1
                    if lowThanma5 >= 3:
                        r.sell(i + 1 + day, df.loc[i + 1 + day, "close"])
                        list_BuyorSell.append(r.Out())
                        Is_selled = True
                        break
 
                # 若卖出条件未出现，则持仓等候
                a = 1+1 # 摸鱼
                temp = temp + 1
 
            if Is_selled == True:
                continue
 
            if i+31 >= len(df)-1: # 检查数组越界
                r.sell(len(df)-1,r.buyPrice)
                list_BuyorSell.append(r.Out())
                break
 
            # 如果最大持股时间达到时 还持有仓位，则强制卖出
            if Is_selled != True:
                r.sell(i + 31,df.loc[i + 31, "close"])
                list_BuyorSell.append(r.Out())
                Is_selled = True
 
            # 结束一个周期
            count = count +1
 
        self.list_BuyorSell = list_BuyorSell
        # print(list_BuyorSell)
 
    # 绘制MA指标图，返回line_ma
    def get_Pyecharts_MA(self,n,index,itheme="light"):
        df = self.df
        colorlist = ["rgb(47,79,79)","rgb(255,140,0)","rgb(0,191,255)","rgb(187, 102, 255)"]
        icolor = colorlist[index-2]
        line = (
            Line(init_opts=opts.InitOpts(theme=itheme,animation_opts=opts.AnimationOpts(animation=False),))
                # 添加x轴交易日期数据
                .add_xaxis(df["trade_date"].tolist())
                .add_yaxis("MA{}".format(n),df["MA{}".format(n)].tolist(),xaxis_index=index,yaxis_index=index,
                           label_opts=opts.LabelOpts(is_show=False),
                           is_symbol_show=False, # 是否显示小圆点
                           itemstyle_opts=opts.ItemStyleOpts(color=icolor)) # 更改颜色
                .set_global_opts(
                    xaxis_opts=opts.AxisOpts(is_show=False,grid_index=index),
                    yaxis_opts=opts.AxisOpts(is_show=False,grid_index=index),
                    legend_opts=opts.LegendOpts(is_show=True),  # 图例是否显示
                )
            )
 
        return line
 
    # 绘制成交量均值线
    def get_Pyecharts_VolMA(self,n,index,itheme="light"):
        df = self.df
        colorlist = ["rgb(47,79,79)","rgb(255,140,0)"]
        icolor = colorlist[index-6]
        line = (
            Line(init_opts=opts.InitOpts(theme=itheme,animation_opts=opts.AnimationOpts(animation=False),))
                # 添加x轴交易日期数据
                .add_xaxis(df["trade_date"].tolist())
                .add_yaxis("VolMA{}".format(n),df["VolMA{}".format(n)].tolist(),xaxis_index=index,yaxis_index=index,
                           label_opts=opts.LabelOpts(is_show=False),
                           is_symbol_show=False, # 是否显示小圆点
                           tooltip_opts=opts.TooltipOpts(is_show=False),
                           itemstyle_opts=opts.ItemStyleOpts(color=icolor)) # 更改颜色
                .set_global_opts(
                    xaxis_opts=opts.AxisOpts(is_show=False,grid_index=index),
                    yaxis_opts=opts.AxisOpts(is_show=False,grid_index=index),
                      # 图例是否显示
                )
            )
 
        return line
 
    # 绘制K线图,
    def get_Pyecharts_Kline(self,itheme="light"):
        tradeAction = [] # 交易输出记录
        df = self.df
        valueList = []
        temp_data = [] # 用于标记区域
        temp_data2 = [] # 用于标记线
        totalProfit = 100
        # 构建日交易金额数据list
        for i in range(len(self.list_BuyorSell)):
            profit = self.list_BuyorSell[i][4] # 0.05
            temp_data.append(opts.MarkAreaItem(name="", x=(df.loc[self.list_BuyorSell[i][0],"trade_date"], df.loc[self.list_BuyorSell[i][1],"trade_date"])))
            temp_data2.append([ { "xAxis": df.loc[self.list_BuyorSell[i][0],"trade_date"],
                                  "yAxis": df.loc[self.list_BuyorSell[i][0],"open"],
                                  "value": "盈利：{:.2f}%".format(profit*100)
                                  },
                                {
                                "xAxis": df.loc[self.list_BuyorSell[i][1], "trade_date"],
                                "yAxis": self.list_BuyorSell[i][3],
                                }])
            totalProfit = totalProfit*(1+profit)
 
            # 输出Excel交易操作信息
            tradeAction.append([ df.loc[self.list_BuyorSell[i][0],"trade_date"], df.loc[self.list_BuyorSell[i][1], "trade_date"], round(df.loc[self.list_BuyorSell[i][0],"open"],2),round(self.list_BuyorSell[i][3],2), "{:.2f}%".format(self.list_BuyorSell[i][4]*100)])
 
        self.totalProfit = totalProfit
            # name="交易{:0>2d}\n{:.2f}%".format(i+1,profit)
        returnList.append([self.name, self.code, tradeAction, round((self.totalProfit-100)/100*100,2)])
 
        for i in range(len(df)):
            valueList.append([df.loc[i, "open"], df.loc[i, "close"], df.loc[i, "high"], df.loc[i, "low"],(df.loc[i,"close"]-df.loc[i,"open"])/df.loc[i,"open"]  ])
        x = df["trade_date"].tolist()
        y = valueList
        # 绘制K线图
        kline = (
            Kline(init_opts=opts.InitOpts(theme=itheme,animation_opts=opts.AnimationOpts(animation=True,animation_easing_update="backOut")))  # chalk
 
                # 添加x轴交易日期数据
                .add_xaxis(x)
                # 添加y轴成交价格数据
                .add_yaxis(series_name="Daily Trade Data", y_axis=y, itemstyle_opts=opts.ItemStyleOpts(  # 风格设置
                    color="red",color0="green",
                    border_color="#ef232a",border_color0="#14b143", # 边框色彩
                    ),
                   markline_opts=opts.MarkLineOpts(
                       label_opts=opts.LabelOpts(
                           position="middle", color="blue"
                       ),
                       data=temp_data2,
                       symbol=["none", "arrow"],
                       symbol_size=10,
                       linestyle_opts=opts.LineStyleOpts(color="blue",width=3)
                   ),
                )
 
                # 设计标记区 x = 开始,结束
                .set_series_opts(markarea_opts=opts.MarkAreaOpts(data=temp_data,
                                                                 itemstyle_opts=opts.ItemStyleOpts(opacity=0.6, # 透明度
                                                                                                color={
                                                                                                   "type":'linear',
                                                                                                   "x":1,"y":1,"x2":0,"y2":0,
                                                                                                   "colorStops": [{"offset": 0, "color": '#F55555'}, {"offset": 1, "color": '#FCCF31'}]
                                                                                                    }
                                                                                                )
                                                                 )
                                 )
                # 设置x、y轴显示信息
                .set_global_opts(xaxis_opts=opts.AxisOpts(name='交易时间'))
                # .set_global_opts(yaxis_opts=opts.AxisOpts(name='交易价格/元'))
                # 固定y轴的范围
                # .set_global_opts(yaxis_opts=opts.AxisOpts(min_=5, max_=10))
 
                .set_global_opts(title_opts=opts.TitleOpts(title="{} {} 总盈利{:.2f}% {}".format(self.code,self.name,(self.totalProfit-100)/100*100,self.isTodayBuy), subtitle='日K线图 交易价格/元'),  # 标题选项
 
                                 legend_opts=opts.LegendOpts(is_show=True),  # 图例选项
 
                                 datazoom_opts=[  # 缩放选项
                                     opts.DataZoomOpts(
                                         is_show=False,
                                         type_="inside",
                                         xaxis_index=[0,1],
                                         # 初始的框选范围
                                         range_start=80,
                                         range_end=100,
                                     ),
                                     opts.DataZoomOpts(
                                         is_show=True,
                                         xaxis_index=[0,1],
                                         type_="slider",
                                         pos_top="95%",
                                         range_start=80,
                                         range_end=100,
                                     )
                                 ],
 
                                 yaxis_opts=opts.AxisOpts(  # 坐标轴配置项
                                     is_scale=True,  # 是否显示y轴
                                     splitarea_opts=opts.SplitAreaOpts(  # 分割区域配置项
                                         is_show=True, areastyle_opts=opts.AreaStyleOpts(opacity=1))
                                 ),
 
                                 # 控制x轴label
                                 xaxis_opts=opts.AxisOpts(axislabel_opts=opts.LabelOpts(formatter=JsCode(
                                     """
                                     function (x) {
                                     a = x.substring(0,4);
                                     b = x.substring(4,6);
                                     c = x.substring(6,8);
                                     return a+'年'+b+'月'+c+'日'; 
                                     }
                                     """)
                                )),
 
 
 
                                 tooltip_opts=opts.TooltipOpts(  # 提示框配置项
                                     trigger="axis",  # 触发类型默认
                                     axis_pointer_type="cross",  # 鼠标指针指示器
                                     background_color="rgba(245, 245, 245, 0.8)",  # 提示框漂浮层背景色
                                     border_width=1,  # 提示框漂浮层边框
                                     border_color="#ccc",
                                     textstyle_opts=opts.TextStyleOpts(color="#000"),  # 提示框文字选项
                                     formatter=JsCode(
                                         """
                                         function (x) {
                                         date = x[0].axisValue.substring(0,4)+ '年' + x[0].axisValue.substring(4,6)+ '月' +x[0].axisValue.substring(6,8)+ '日';
                                         open = x[0].data[1];
                                         close = x[0].data[2];
 
                                         
                                         return date + '<br>' + '开盘价：' + open + '<br>' +'收盘价：' + close + '<br>' +'涨跌幅：' + Math.round((close-open)/close*100*100)/100 + '%<br>'+ x[1].seriesName +'&nbsp;&nbsp;：'+ x[1].data[1] + '<br>' + x[2].seriesName +'：'+ x[2].data[1] + '<br>'+ x[3].seriesName +'：'+ x[3].data[1] + '<br>'+ x[4].seriesName +'：'+ x[4].data[1] + '<br>'; 
                                         }
                                         """)
                                 ),
 
                                 axispointer_opts=opts.AxisPointerOpts(  # 坐标轴指示器配置项
                                     is_show=True,
                                     label=opts.LabelOpts(background_color="#777"),
                                 ),
 
                                 brush_opts=opts.BrushOpts(  # 区域选择组建配置项
                                     x_axis_index="all",  # 指定哪些 xAxisIndex 可以被刷选
                                     brush_link="all",  # 不同系列间，选中的项可以联动。
                                     out_of_brush={"colorAlpha": 0.1},
                                     brush_type="lineX",
                                 ),
                                )
        )
        return kline
 
    # 绘制成交量图形
    def get_Pyecharts_Bar(self,itheme="light"):
        df = self.df
        valueList = []
        # 构建日交易金额数据list
        for i in range(len(df)):
            valueList.append([df.loc[i, "open"], df.loc[i, "close"], df.loc[i, "high"], df.loc[i, "low"]])
        # 绘制成交量柱状图
        bar = (
            Bar()
                .add_xaxis(xaxis_data=df["trade_date"].tolist())
                .add_yaxis(series_name="Volume",y_axis=df["vol"].tolist(),label_opts=opts.LabelOpts(is_show=False),
                    # 设置多图联动
                    xaxis_index=1,
                    yaxis_index=1,
                    tooltip_opts=opts.TooltipOpts(is_show=False),)
 
                .set_global_opts(
                    xaxis_opts=opts.AxisOpts(
 
                        # 控制x轴label
                        axislabel_opts=opts.LabelOpts(formatter=JsCode(
                            """
                            function (x) {
                            a = x.substring(0,4);
                            b = x.substring(4,6);
                            c = x.substring(6,8);
                            return a+'年'+b+'月'+c+'日'; 
                            }
                            """)
                        ),
                        type_="category",
                        is_scale=True,
                        grid_index=1,
                        axisline_opts=opts.AxisLineOpts(is_on_zero=True),
                        axistick_opts=opts.AxisTickOpts(is_show=True),
                        splitline_opts=opts.SplitLineOpts(is_show=False),
                        split_number=20,
                        min_="dataMin",
                        max_="dataMax",),
 
                    yaxis_opts=opts.AxisOpts(
                        grid_index=1,
                        is_scale=True,
                        split_number=2,
                        axislabel_opts=opts.LabelOpts(is_show=True),
                        axisline_opts=opts.AxisLineOpts(is_show=True),
                        axistick_opts=opts.AxisTickOpts(is_show=True),
                        splitline_opts=opts.SplitLineOpts(is_show=False),),
                legend_opts=opts.LegendOpts(is_show=False),
            )
        )
        return bar
 
    # 绘制主图并输出页面
    def Print_Main_index(self,kline,bar_volumn,line_ma=None,line_ma2=None,line_ma3=None,line_ma4 = None,itheme="light"):
        bar = bar_volumn
 
        kline.overlap(line_ma)
        kline.overlap(line_ma2)
        kline.overlap(line_ma3)
        kline.overlap(line_ma4)
 
        grid_chart = Grid(
            init_opts=opts.InitOpts(
                width="1200px", height="580px",
                animation_opts=opts.AnimationOpts(animation=True,animation_easing="linear"),
                theme=itheme, page_title="Pyecharts_Demo",
            )
        )
        # 添加上图
        grid_chart.add(
            kline,
            grid_opts=opts.GridOpts(pos_left="10%", pos_right="10%", height="60%"),
        )
        # 添加下图
        grid_chart.add(
            bar,
            grid_opts=opts.GridOpts(pos_left="10%", pos_right="10%", pos_top="75%", height="16%"),
        )
        #print("已保存文件至{}".format("{}-{}.html".format(self.code[0:6], self.name)))
        grid_chart.render(path = "./Django/Web/user_manage/templates/{}.html".format(self.code[0:6]))
        import os
        print("已保存文件至{}".format("{}-{}.html".format(self.code[0:6],self.name)))
        # os.system("start C:/Users/***/Desktop/pyecharts-{}-{}.html".format(self.code[0:6],self.name))
 
    def begin(self):
        self.get_daily_trade()
        # self.get_Macd()
        self.get_MA(5)
        self.get_MA(10)
        self.get_MA(20)
        self.get_MA(150)
        self.get_Vol_MA(5)
        self.get_Vol_MA(10)
        self.set_Trading_Strategy()
        self.line_ma5 = self.get_Pyecharts_MA(5,2) # index由2开始
        self.line_ma10 = self.get_Pyecharts_MA(10,3)
        self.line_ma20 = self.get_Pyecharts_MA(20,4)
        self.line_ma50 = self.get_Pyecharts_MA(150,5)
        self.line_volma5 = self.get_Pyecharts_VolMA(5, 6)
        self.line_volma10 = self.get_Pyecharts_VolMA(10, 7)
        self.kline = self.get_Pyecharts_Kline()
        self.bar_volumn = self.get_Pyecharts_Bar().overlap(self.line_volma5).overlap(self.line_volma10)
        self.Print_Main_index(self.kline, self.bar_volumn, self.line_ma5, self.line_ma10, self.line_ma20, self.line_ma50)
 
class ExcelWriter:
    path = None
    wb = None
    ws = None
 
    def setLocation(self,Path):
        self.path = Path
 
    def newFile(self):
        wb = Workbook()
        ws = wb.active
        self.wb = wb
        self.ws = ws
 
    def outList(self):
        list = []
        a = -100
        size = 5
        for i in range(70):
            a = a
            b = a + size
            list.append([[a, b], "{}%->{}%".format(a, b), 0])
            a = a + size
        return list
 
    def inList(self,n,list):
        n = int(n)
        for i in list:
            # print(type(n), type(i[0][0]))
            if int(i[0][0]) <= n and n < int(i[0][1]):
                i[2] = i[2] + 1
        return list
 
    def writeData(self,Lists): # List 单个元素内容 [stockName,stockCode,[操作1,操作2,...],总盈利]
        dictProfit = self.outList() # 收益率计数列表
        countColumn = 1 # 控制列的输出
        countRow = 1 # 控制行的输出
        Avprofit = 0
        for stock in Lists:
            countRow = 1
            # 输出股票名称及代码
            self.ws.cell(row=countRow,column=countColumn).value = stock[0] + stock[1]
            countRow = countRow+1
 
            # 输出操作信息
            for action in stock[2]:
                self.ws.cell(row=countRow, column=countColumn).value = "{} {} {} {} {}".format(action[0],action[1],action[2],action[3],action[4])
                countRow = countRow+1
 
            self.ws.cell(row=25, column=countColumn).value = "个股收益："
            self.ws.cell(row=26, column=countColumn).value = str(stock[3]) + "%"
            dictProfit = self.inList(stock[3],dictProfit)
            Avprofit = Avprofit + stock[3]
            countColumn = countColumn + 1
        # print(profit)
        self.ws.cell(row=27, column=1).value = "平均收益："
        self.ws.cell(row=28, column=1).value = str(round(Avprofit/len(Lists),2)) + "%"
 
        self.ws.cell(row=30, column=1).value = "收益分布分析："
        countColumn2 = 1
        for i in dictProfit:
            self.ws.cell(row=31, column=countColumn2).value = i[2]
            self.ws.cell(row=32, column=countColumn2).value = i[1]
            countColumn2 = countColumn2 + 1
 
 
    def saveExcel(self):
        print("保存文件至 -> {}".format(self.path))
        self.wb.save(self.path)
        #import os
        #os.system("start  C:/Users/***/Desktop/StockAnalysis/totalAnalysis.xlsx")
 
 
 
 
 
if __name__ == '__main__':
    returnList = []
    #codeList = ["000615"]
    codeList = ["000615","000628","000629","000635","000659","000663","000665","000666","000670","000679","000680"]
    #with open("C:/Users/***/Desktop/StockAnalysis/code.txt", 'r') as file:
    #    data = file.readlines()
    #    for i in data:
    #        codeList.append(i.replace("\n",""))
    #print(codeList)
 
    # for i in range(10):
    #     code = input("请输入证券代码：")
    #     codeList.append(code)
 
    for i in codeList:
        stock1 = Stock()
        stock1.set_code(i)
        stock1.start = "20000101"
        stock1.end = "20230818"
        stock1.begin()
    # ew = ExcelWriter()
    # ew.setLocation("./totalAnalysis.xlsx")
    # ew.newFile()
    # ew.writeData(returnList)
    # ew.saveExcel()